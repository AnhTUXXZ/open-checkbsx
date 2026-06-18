import sys
import os
import time
import sqlite3
import threading
import socket
import io
import re
import base64
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox

from flask import Flask, request, jsonify, render_template
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import openpyxl

import cv2
import numpy as np

# ==========================================
# 1. CẤU HÌNH HỆ THỐNG
# ==========================================

DATA_DIR = 'data'
TEMP_IMAGES_DIR = 'temp_images'
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(TEMP_IMAGES_DIR, exist_ok=True)

DB_PARKING = os.path.join(DATA_DIR, 'parking.db')
pytesseract.pytesseract.tesseract_cmd = os.path.join(DATA_DIR, 'Tesseract-OCR', 'tesseract.exe')

app = Flask(__name__)
app.secret_key = 'parking_admin_secret_key'
sheet_lock = threading.Lock()

EXCEL_FILE = 'thong_ke_bai_xe.xlsx'

def init_db():
    conn = sqlite3.connect(DB_PARKING)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS parking_logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                 plate TEXT, 
                 time_in DATETIME, 
                 time_out DATETIME, 
                 status TEXT)''')
    conn.commit()
    conn.close()

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LỊCH SỬ RA VÀO"
        ws.append(["ID", "Biển Số", "Thời Gian Vào", "Thời Gian Ra", "Trạng Thái"])
        wb.save(EXCEL_FILE)

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

# ==========================================
# 2. FLASK ROUTES (WEB NHẬN ẢNH)
# ==========================================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/ping', methods=['POST'])
def ping():
    return jsonify({'status': 'ok'})

@app.route('/upload-ocr', methods=['POST'])
def process_image():
    if 'image' not in request.files:
        return jsonify({'error': 'Không tìm thấy ảnh'}), 400
    
    file = request.files['image']
    try:
        in_memory_file = io.BytesIO(file.read())
        file_bytes = np.frombuffer(in_memory_file.getvalue(), dtype=np.uint8)
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)

        # Tiền xử lý ảnh chuyên sâu
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        
        sharpen_kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
        sharpened = cv2.filter2D(gray, -1, sharpen_kernel)

        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        clahe_img = clahe.apply(sharpened)

        _, thresh = cv2.threshold(cv2.GaussianBlur(clahe_img, (5, 5), 0), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        # Đóng chữ (Nối các nét bị đứt do mờ)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
        morph = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
        
        # MỚI: Mở chữ (Tẩy các vết ốc vít, chấm đen nhỏ để không đọc nhầm 0 thành 9)
        kernel_small = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
        open_morph = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel_small)

        # Đưa thêm open_morph vào đầu mảng ưu tiên để AI quét trước
        filters = [open_morph, morph, thresh, clahe_img, sharpened, gray]
        psms = [11, 7, 6, 3, 12] 
        
        plate_text = ""
        best_d = None
        image_base64 = ""

        for f_img in filters:
            for psm in psms:
                custom_config = f'--oem 3 --psm {psm}'
                raw_text = pytesseract.image_to_string(f_img, config=custom_config)
                
                clean_text = ''.join(e for e in raw_text if e.isalnum()).upper()
                matches = re.findall(r'[1-9]\d[A-Z][A-Z0-9]?\d{4,5}', clean_text)
                
                if matches:
                    plate_text = max(matches, key=len)
                    best_d = pytesseract.image_to_data(f_img, config=custom_config, output_type=pytesseract.Output.DICT)
                    break 
            if plate_text:
                break 

        if plate_text:
            try:
                box_x_min, box_y_min, box_x_max, box_y_max = 9999, 9999, 0, 0
                for i in range(len(best_d['text'])):
                    word = ''.join(e for e in best_d['text'][i] if e.isalnum()).upper()
                    if word and (word in plate_text):
                        x = int(best_d['left'][i] / 2)
                        y = int(best_d['top'][i] / 2)
                        w = int(best_d['width'][i] / 2)
                        h = int(best_d['height'][i] / 2)
                        
                        box_x_min = min(box_x_min, x)
                        box_y_min = min(box_y_min, y)
                        box_x_max = max(box_x_max, x + w)
                        box_y_max = max(box_y_max, y + h)

                if box_x_min != 9999:
                    cv2.rectangle(img, (box_x_min - 15, box_y_min - 15), (box_x_max + 15, box_y_max + 15), (0, 255, 0), 3)
                    cv2.putText(img, plate_text, (box_x_min - 15, box_y_min - 20), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                
                _, buffer = cv2.imencode('.jpg', img)
                image_base64 = base64.b64encode(buffer).decode('utf-8')
            except Exception as e:
                pass
        else:
            return jsonify({'error': 'Ảnh có quá nhiều cảnh vật dư thừa hoặc bị lóa. Vui lòng căn biển số lọt vào khung ngắm xanh lá!'}), 400

        # --- LOGIC RA/VÀO & TÍNH TIỀN ---
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = "VÔ"
        time_in = current_time
        time_out = ""
        duration_str = ""
        fee = 0

        conn = sqlite3.connect(DB_PARKING)
        c = conn.cursor()
        
        c.execute("SELECT id, time_in FROM parking_logs WHERE plate=? AND status='VÔ' ORDER BY id DESC LIMIT 1", (plate_text,))
        row = c.fetchone()
        
        if row:
            log_id, db_time_in = row
            status = "RA"
            time_in = db_time_in
            time_out = current_time
            c.execute("UPDATE parking_logs SET time_out=?, status=? WHERE id=?", (time_out, status, log_id))
            
            try:
                t_in = datetime.strptime(time_in, "%Y-%m-%d %H:%M:%S")
                t_out = datetime.strptime(time_out, "%Y-%m-%d %H:%M:%S")
                delta = t_out - t_in
                minutes = int(delta.total_seconds() // 60)
                
                fee = minutes * 50
                duration_str = f"{minutes} phút"
                msg = f"Đậu {duration_str}. Thu {fee:,}đ."
            except Exception:
                msg = "Xe xuất bến thành công."
        else:
            c.execute("INSERT INTO parking_logs (plate, time_in, time_out, status) VALUES (?, ?, ?, ?)", 
                      (plate_text, time_in, "", status))
            log_id = c.lastrowid
            msg = "Đã lưu thông tin xe vào bãi."
            
        conn.commit()
        conn.close()

        with sheet_lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            if status == "VÔ":
                ws.append([log_id, plate_text, time_in, "", status])
            else:
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value == log_id:
                        ws.cell(row=r, column=4, value=time_out)
                        ws.cell(row=r, column=5, value=status)
                        break
            wb.save(EXCEL_FILE)

        return jsonify({
            'plate': plate_text,
            'status': status,
            'time': current_time,
            'message': msg,
            'image_base64': image_base64,
            'duration': duration_str,
            'fee': fee
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

def get_stats():
    try:
        conn = sqlite3.connect(DB_PARKING)
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM parking_logs WHERE status='VÔ'")
        inside = c.fetchone()[0]
        
        today_str = datetime.now().strftime("%Y-%m-%d")
        c.execute("SELECT COUNT(*) FROM parking_logs WHERE status='RA' AND time_out LIKE ?", (f"{today_str}%",))
        out_today = c.fetchone()[0]
        conn.close()
        return inside, out_today
    except:
        return 0, 0

def get_recent_logs(limit=50): 
    try:
        conn = sqlite3.connect(DB_PARKING)
        c = conn.cursor()
        c.execute("SELECT plate, time_in, time_out, status FROM parking_logs ORDER BY id DESC LIMIT ?", (limit,))
        rows = c.fetchall()
        conn.close()
        return rows
    except:
        return []

# ==========================================
# 3. GIAO DIỆN DESKTOP (TKINTER)
# ==========================================

class DesktopParkingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Admin - Quản Lý Bãi Giữ Xe")
        self.root.geometry("1100x650") 
        self.root.configure(bg="#f1f5f9") 
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", font=('Inter', 11, 'bold'), background="#e2e8f0", foreground="#1e293b")
        style.configure("Treeview", font=('Inter', 10), rowheight=30, background="white", fieldbackground="white")
        style.map('Treeview', background=[('selected', '#eff6ff')], foreground=[('selected', '#1d4ed8')])

        self.create_dashboard()
            
    def create_dashboard(self):
        header_card = tk.Frame(self.root, bg="white", bd=1, relief="solid")
        header_card.pack(fill='x', padx=20, pady=15)
        
        ip_address = get_local_ip()
        tk.Label(header_card, text=f"🟢 Server Web Đang Chạy: http://{ip_address}:5000", 
                 font=('Inter', 13, 'bold'), fg='#059669', bg='white').pack(pady=10)
        
        cards_frame = tk.Frame(self.root, bg="#f1f5f9")
        cards_frame.pack(fill='x', padx=20, pady=(0, 15))
        
        card1 = tk.Frame(cards_frame, bg="white", bd=1, relief="solid")
        card1.pack(side='left', expand=True, fill='both', padx=(0, 10))
        tk.Frame(card1, bg="#ef4444", height=5).pack(fill='x', side='top')
        tk.Label(card1, text="XE ĐANG TRONG BÃI", font=('Inter', 12, 'bold'), fg='#64748b', bg='white').pack(pady=(15,5))
        self.lbl_inside = tk.Label(card1, text="0", font=('Inter', 36, 'bold'), fg='#1e293b', bg='white')
        self.lbl_inside.pack(pady=(0, 15))

        card2 = tk.Frame(cards_frame, bg="white", bd=1, relief="solid")
        card2.pack(side='left', expand=True, fill='both', padx=(10, 0))
        tk.Frame(card2, bg="#10b981", height=5).pack(fill='x', side='top')
        tk.Label(card2, text="LƯỢT XE RA (HÔM NAY)", font=('Inter', 12, 'bold'), fg='#64748b', bg='white').pack(pady=(15,5))
        self.lbl_out = tk.Label(card2, text="0", font=('Inter', 36, 'bold'), fg='#1e293b', bg='white')
        self.lbl_out.pack(pady=(0, 15))

        tables_container = tk.Frame(self.root, bg="#f1f5f9")
        tables_container.pack(expand=True, fill='both', padx=20, pady=(0, 20))

        frame_in = tk.Frame(tables_container, bg="white", bd=1, relief="solid")
        frame_in.pack(side='left', expand=True, fill='both', padx=(0, 10))
        
        tk.Label(frame_in, text="🔴 LỊCH SỬ XE VÀO BÃI", font=('Inter', 11, 'bold'), fg='#ef4444', bg='white').pack(anchor='w', padx=15, pady=10)
        
        self.tree_in = ttk.Treeview(frame_in, columns=("icon", "plate", "time_in"), show='headings', height=10)
        self.tree_in.heading("icon", text="Trạng Thái")
        self.tree_in.heading("plate", text="Biển Số Xe")
        self.tree_in.heading("time_in", text="Thời Gian Vào")
        
        self.tree_in.column("icon", anchor="center", width=90)
        self.tree_in.column("plate", anchor="center", width=120)
        self.tree_in.column("time_in", anchor="center", width=180)
        
        self.tree_in.tag_configure('tag_VO', foreground='#ef4444', font=('Inter', 10, 'bold'))
        self.tree_in.pack(fill='both', expand=True, padx=15, pady=(0, 15))

        frame_out = tk.Frame(tables_container, bg="white", bd=1, relief="solid")
        frame_out.pack(side='left', expand=True, fill='both', padx=(10, 0))
        
        tk.Label(frame_out, text="🟢 XE XUẤT BẾN & TÍNH TIỀN (Tự ẩn sau 30s)", font=('Inter', 11, 'bold'), fg='#10b981', bg='white').pack(anchor='w', padx=15, pady=10)
        
        self.tree_out = ttk.Treeview(frame_out, columns=("icon", "plate", "duration", "fee"), show='headings', height=10)
        self.tree_out.heading("icon", text="Trạng Thái")
        self.tree_out.heading("plate", text="Biển Số")
        self.tree_out.heading("duration", text="Thời Gian Đậu")
        self.tree_out.heading("fee", text="Thành Tiền")
        
        self.tree_out.column("icon", anchor="center", width=90)
        self.tree_out.column("plate", anchor="center", width=120)
        self.tree_out.column("duration", anchor="center", width=110)
        self.tree_out.column("fee", anchor="center", width=130)
        
        self.tree_out.tag_configure('tag_RA', foreground='#10b981', font=('Inter', 10, 'bold'))
        self.tree_out.pack(fill='both', expand=True, padx=15, pady=(0, 15))

        self.update_dashboard()

    def update_dashboard(self):
        inside, out_today = get_stats()
        self.lbl_inside.config(text=str(inside))
        self.lbl_out.config(text=str(out_today))

        for item in self.tree_in.get_children(): self.tree_in.delete(item)
        for item in self.tree_out.get_children(): self.tree_out.delete(item)
            
        logs = get_recent_logs()
        current_time = datetime.now()

        for log in logs:
            plate, time_in, time_out, status = log
            
            if status == "VÔ":
                self.tree_in.insert("", "end", values=("📥 VÀO", plate, time_in), tags=('tag_VO',))
            elif status == "RA" and time_out:
                try:
                    t_in = datetime.strptime(time_in, "%Y-%m-%d %H:%M:%S")
                    t_out = datetime.strptime(time_out, "%Y-%m-%d %H:%M:%S")
                    seconds_diff = (current_time - t_out).total_seconds()
                    
                    if seconds_diff <= 30:
                        minutes = int((t_out - t_in).total_seconds() // 60)
                        fee = minutes * 50
                        duration_str = f"{minutes} phút"
                        fee_str = f"{fee:,} VNĐ"
                        
                        self.tree_out.insert("", "end", values=("📤 RA", plate, duration_str, fee_str), tags=('tag_RA',))
                except Exception:
                    pass 

        self.root.after(2000, self.update_dashboard)

def run_flask_server():
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)

if __name__ == '__main__':
    init_db()
    init_excel()
    
    threading.Thread(target=run_flask_server, daemon=True).start()
    
    root = tk.Tk()
    app_gui = DesktopParkingApp(root)
    root.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()